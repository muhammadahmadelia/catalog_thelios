from email.mime import image
import os
from re import L
import sys
import json
from time import sleep
from urllib.parse import quote
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
import chromedriver_autoinstaller

from models.store import Store
from models.brand import Brand
from models.product import Product
from models.variant import Variant
from models.metafields import Metafields
import glob
import requests
from bs4 import BeautifulSoup
# import pandas as pd
import threading

from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
from openpyxl.utils import get_column_letter
from PIL import Image

from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

class myScrapingThread(threading.Thread):
    def __init__(self, threadID: int, name: str, obj, product_urls: list[str], headers: dict, brand: str, glasses_type: str, frame_code: str) -> None:
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.product_urls = product_urls
        self.brand = brand
        self.glasses_type = glasses_type
        self.headers = headers
        self.frame_code = frame_code
        self.obj = obj
        self.status = 'in progress'
        pass

    def run(self):
        self.obj.scrape_products(self.product_urls, self.headers, self.brand, self.glasses_type, self.frame_code)
        self.status = 'completed'

    def active_threads(self):
        return threading.activeCount()


class Thelios_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str) -> None:
        self.DEBUG = DEBUG
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.args = ["hide_console", ]
        # self.browser = webdriver.Chrome(options=self.chrome_options, service_args=self.args)
        self.browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=self.chrome_options)
        self.data = []
        pass

    def controller(self, store: Store, brands_with_types: list[dict]):
        try:
            cookies = ''

            self.browser.get(store.link)
            self.wait_until_browsing()

            if self.login(store.username, store.password):
                if self.wait_until_element_found(20, 'css_selector', 'div[data-label="Sun"] > a'):
                    print('Scraping products for')
                    for brand_with_type in brands_with_types:
                        brand: str = brand_with_type['brand']

                        for glasses_type in brand_with_type['glasses_type']:
                            
                            print(f'Brand: {brand} | Type: {str(glasses_type).strip().title()}')
                            scraped_products = 0
                            brand_url = ''
                            
                            if str(glasses_type).strip().lower() == 'sunglasses':
                                if str('Celine').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3ASole%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3ACeline&text=&newArrivals=false#'
                                elif str('Dior').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3ASole%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3ADior&text=&newArrivals=false#'
                                elif str('Fendi').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3ASole%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3AFendi&text=&newArrivals=false#'
                                elif str('Givenchy').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3ASole%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3AGivenchy&text=&newArrivals=false#'
                                elif str('Stella McCartney').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3ASole%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3AStella%2BMcCartney&text=&newArrivals=false#'
                            elif str(glasses_type).strip().lower() == 'eyeglasses':
                                if str('Celine').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3AVista%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3ACeline&text=&newArrivals=false#'
                                elif str('Dior').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3AVista%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3ADior&text=&newArrivals=false#'
                                elif str('Fendi').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3AVista%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3AFendi&text=&newArrivals=false#'
                                elif str('Givenchy').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3AVista%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3AGivenchy&text=&newArrivals=false#'
                                elif str('Stella McCartney').strip().lower() == str(brand).strip().lower():
                                    brand_url = 'https://my.thelios.com/it/it/Maison/c/00?q=%3Arelevance%3Atype%3AVista%3AfittingDescription%3AUniversal%3AfittingDescription%3AInternational%3Apurchasable%3Apurchasable%3AimShip%3Afalse%3AnewArrivals%3Afalse%3AallCategoriesForName%3AStella%2BMcCartney&text=&newArrivals=false#'
                            
                            self.browser.get(brand_url)
                            self.wait_until_browsing()

                            while True:

                                self.wait_until_element_found(30, 'css_selector', 'div[class="product__listing product__grid col-xs-12 "] > div')

                                for div_tag in self.browser.find_elements(By.CSS_SELECTOR, 'div[class="product__listing product__grid col-xs-12 "] > div'):
                                    try:
                                        try: ActionChains(self.browser).move_to_element(div_tag).perform()
                                        except: pass

                                        product_urls = []
                            
                                        scraped_products += 1
                                        product_urls = []
                                        for a_tag in div_tag.find_elements(By.CSS_SELECTOR, 'div[class="preview-carousel-variant"] > div > a'):
                                            product_url = str(a_tag.get_attribute('href')).strip()
                                            if 'https://my.thelios.com' not in product_url:
                                                product_url = f'https://my.thelios.com{product_url}'

                                            if product_url not in product_urls: product_urls.append(product_url)
                                        frame_code = str(div_tag.get_attribute('onclick')).strip().split("'")[-2].strip()

                                        if not cookies: cookies = self.get_cookies_from_browser()
                                        headers = self.get_headers(cookies, self.browser.current_url)
                                        
                                        # self.scrape_products(product_urls, headers, brand, glasses_type, frame_code)
                                        self.create_thread(product_urls, headers, brand, glasses_type, frame_code)

                                        if self.thread_counter >= 10: 
                                            self.wait_for_thread_list_to_complete()
                                            self.save_to_json(self.data)
                                    
                                    except Exception as e:
                                        if self.DEBUG: print(f'Exception in product loop: {e}')
                                        else: pass

                                next_page_li = self.browser.find_element(By.CSS_SELECTOR, 'li[class*="pagination-next"]')
                                if next_page_li and 'disabled' not in next_page_li.get_attribute('class'):
                                    href = next_page_li.find_element(By.TAG_NAME, 'a').get_attribute('href')
                                    if 'https://my.thelios.com' not in href: href = f'https://my.thelios.com{href}'
                                    self.browser.get(href)
                                    self.wait_until_browsing()
                                else: break 

                            self.wait_for_thread_list_to_complete()
                            self.save_to_json(self.data)
            else: print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in scraper controller: {e}')
            else: pass
        finally: 
            self.wait_for_thread_list_to_complete()
            self.save_to_json(self.data)
            self.browser.quit()

    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.2)
            except: pass

    def login(self, username: str, password: str) -> bool:
        login_flag = False
        try:
            if self.wait_until_element_found(20, 'xpath', '//input[@id="j_username"]'):
                self.browser.find_element(By.XPATH, '//input[@id="j_username"]').send_keys(username)
                self.browser.find_element(By.XPATH, '//input[@id="j_password"]').send_keys(password)
                try:
                    button = WebDriverWait(self.browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[class*="btn-block"]')))
                    button.click()

                    WebDriverWait(self.browser, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[data-label="Sun"] > a')))
                    login_flag = True
                except Exception as e: print(e)
        except Exception as e:
            if self.DEBUG: print(f'Exception in login: {str(e)}')
            else: pass
        finally: return login_flag

    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    def open_new_tab(self, url: str) -> None:
        # open category in new tab
        self.browser.execute_script('window.open("'+str(url)+'","_blank");')
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
        self.wait_until_browsing()
    
    def close_last_tab(self) -> None:
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
    
    def get_cookies_from_browser(self) -> str:
        cookies = ''
        try:
            browser_cookies = self.browser.get_cookies()
        
            for browser_cookie in browser_cookies:
                if browser_cookie['name'] == '_hjIncludedInPageviewSample': 
                    cookies = f'{browser_cookie["name"]}={browser_cookie["value"]}; __utmt=1; {cookies}'
                else: cookies = f'{browser_cookie["name"]}={browser_cookie["value"]}; {cookies}'
            cookies = cookies.strip()[:-1]
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_cookies_from_browser: {e}')
            self.print_logs(f'Exception in get_cookies_from_browser: {e}')
        finally: return cookies

    def get_headers(self, cookies: str, referer: str):
        return {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'en-US,en;q=0.9',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Cookie': cookies,
            'Host': 'my.thelios.com',
            'Referer': referer,
            'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36'
        }

    def scrape_products(self, product_urls: list[str], headers: dict, brand: str, glasses_type: str, frame_code: str):
        try:
            for product_url in product_urls:
                response = self.make_request(product_url, headers)
                if response:
                    soup = BeautifulSoup(response.text, 'lxml')

                    product = Product()
                    product.brand = brand
                    product.url = product_url
                    product.type = glasses_type

                    product.frame_code = frame_code
                    
                    try: product.lens_code = str(soup.select_one('span[class="productColour-pdp"]').text).strip()
                    except: pass

                    product.number = f'{product.frame_code} {product.lens_code}'
                    
                    try:
                        text = str(soup.select_one('div[class$="landscape-pdp-space"] > div').text).strip()
                        product.frame_color = str(text).split(',')[0].strip().title().replace('\u00a0', ' ')
                        product.lens_color = str(text).split(',')[-1].strip().title().replace('\u00a0', ' ')
                    except: pass

                    if not str(product.number): product.number = f'{product.frame_code} {product.lens_code}'

                    variant = Variant()
                    try: variant.title = str(soup.select_one('div[class="variant-selector"] > div[class*="col-md-12"] > a > button').text).strip()
                    except: pass
                    try: variant.sku = f'{product.number} {variant.title}'
                    except: pass
                    variant.found_status = 1
                    try:
                        p_tag_class = soup.select_one('p[class*="stock-status"]').get('class')
                        if 'instock' in p_tag_class: variant.inventory_quantity = 1
                        else: variant.inventory_quantity = 0
                    except: pass

                    try:
                        for price_div_tag in soup.select('div[class="price-box"]'):
                            variant.listing_price = str(price_div_tag.text).strip().replace('€', '').strip()
                            if variant.listing_price:
                                if variant.listing_price[-3:] == ',00': 
                                    variant.listing_price = f'{variant.listing_price[:-3]}.00'
                    except: pass
                    
                    metafields = Metafields()
                    try: 
                        metafields.img_url = str(soup.select_one('div[class="carousel image-gallery__image js-gallery-image"] > div[class="item"] > img[class="lazyOwl"]').get('data-zoom-image')).strip()
                        if 'https://my.thelios.com' not in metafields.img_url: metafields.img_url = f'https://my.thelios.com{metafields.img_url}' 
                    except: pass

                    try:
                        s = requests.session()
                        try: s.patch(url='https://my.thelios.com/it/it/my-account/update-selected-price', params={'PRIVATE': ''}, headers=headers)
                        except Exception as e: self.print_logs(e)
                        response3 = s.get(url=product_url, headers=headers)
                        
                        if response3.status_code == 200:
                            soup2 = BeautifulSoup(response3.text, 'lxml')
                            
                            for price_div_tag in soup2.select('div[class="price-box"]'):
                                variant.wholesale_price = str(price_div_tag.text).strip().replace('€', '').strip()
                                if variant.wholesale_price:
                                    if variant.wholesale_price[-3:] == ',00': 
                                        variant.wholesale_price = f'{variant.wholesale_price[:-3]}.00'
                                            
                    except Exception as e: self.print_logs(e)

                    product.metafields = metafields
                    product.variants = variant
                    self.data.append(product)
        
        except Exception as e:
            self.print_logs(f'Exception in scrape_products: {e}')
            if self.DEBUG: print(f'Exception in scrape_products: {e}')

    def make_request(self, url: str, headers: dict):
        response = None
        for _ in range(0, 10):
            try:
                response = requests.get(url=url, headers=headers)
                if response == 200: break
                else: sleep(0.5)
            except Exception as e:
                self.print_logs(f'Exception in make_request: {e}')
                if self.DEBUG: print(f'Exception in make_request: {e}')
                sleep(0.5)
        return response

    def create_thread(self, product_urls: list[str], headers: dict, brand: str, glasses_type: str, frame_code: str) -> None:
        thread_name = "Thread-"+str(self.thread_counter)
        self.thread_list.append(myScrapingThread(self.thread_counter, thread_name, self, product_urls, headers,  brand, glasses_type, frame_code))
        self.thread_list[self.thread_counter].start()
        self.thread_counter += 1

    def is_thread_list_complted(self) -> bool:
        for obj in self.thread_list:
            if obj.status == "in progress":
                return False
        return True

    def wait_for_thread_list_to_complete(self) -> None:
        while True:
            result = self.is_thread_list_complted()
            if result: 
                self.thread_counter = 0
                self.thread_list.clear()
                break
            else: sleep(1)

    def click_to_make_price_visible(self) -> None:
        try:
            if not self.is_price_visible():
                element = self.browser.find_element(By.XPATH, '//svg-icon[@class="eye-icon"]')
                ActionChains(self.browser).move_to_element(element).click().perform()
                sleep(0.4)

                for li in self.browser.find_elements(By.XPATH, '//ul[@aria-labelledby="basic-link"]/li'):
                    if str('Cost and SRP').strip().lower() in str(li.text).strip().lower():
                        ActionChains(self.browser).move_to_element(li).click().perform()
                        self.wait_until_price_is_shown()
        except Exception as e:
            if self.DEBUG: print(f'Exception in click_to_make_price_visible: {str(e)}')
            else: pass

    def is_price_visible(self) -> bool:
        try:
            for td in self.browser.find_elements(By.XPATH, '//table[@class="table table-borderless"]/tbody/tr/td'):
                if str('Suggested Retail Price').strip().lower() in str(td.text).strip().lower(): return True
            return False
        except: return False

    def wait_until_price_is_shown(self) -> bool:
        flag = False
        for _ in range(0, 30):
            try:
                tds_label = self.browser.find_elements(By.XPATH, '//table[@class="table table-borderless"]/tbody/tr/td')
                for i in range(0, len(tds_label)):
                    if str('Suggested Retail Price').strip().lower() in str(tds_label[i].text).strip().lower(): 
                        flag = True
                        break
                if flag: break
            except: sleep(0.3)
            finally: return flag

    def get_size_price_status(self):
        size_titles, wholesale_prices, listing_prices, availability, gtin = [], [], [], [], []
        sizes = []
        caliber, rod, bridge = '', '', ''
        try:
            trs = self.browser.find_elements(By.XPATH, '//table[@class="table table-borderless"]/tbody/tr')
            
            for j in range(1, len(trs)):
                tr = trs[j].find_element(By.XPATH, '//table[@class="table table-borderless inner-table"]/tr')
                
                # tds_value = tr.find_elements_by_tag_name('td')
                try:
                    value = str(tr.find_elements(By.XPATH, ".//td[contains(text(), '€')]")[0].text).replace('€', '').strip()
                    value = f"{str(value[0:-3]).strip().replace(',', '').strip()}.00"
                    wholesale_prices.append(value)
                except Exception as e: 
                    if self.DEBUG: print(f'Exception in wholesale_price: {str(e)}')
                    else: pass

                try:
                    value = str(tr.find_elements(By.XPATH, ".//td[contains(text(), '€')]")[1].text).replace('€', '').strip()
                    value = f"{str(value[0:-3]).strip().replace(',', '').strip()}.00"
                    listing_prices.append(value)
                except Exception as e: 
                    if self.DEBUG: print(f'Exception in listing_prices: {str(e)}')
                    else: pass
                
            for tr in self.browser.find_elements(By.XPATH, '//table[@class="table table-borderless inner-table"]/tr'):
                tds_value = tr.find_elements(By.TAG_NAME, 'td')
                try:
                    caliber, rod, bridge = str(tds_value[0].text).strip(), str(tds_value[1].text).strip(), str(tds_value[2].text).strip()
                    # print(len(tds_value), f'{caliber}-{rod}-{bridge}')
                    size_titles.append(caliber)
                    sizes.append(f'{caliber}-{rod}-{bridge}')
                    # if not metafields.product_size: metafields.product_size =f'{caliber}-{rod}-{bridge}'
                    caliber, rod, bridge = '', '', ''
                except Exception as e: 
                    if self.DEBUG: print(f'Exception in product size: {str(e)}')
                    else: pass

                
                try:
                    span_tag_class = tr.find_element(By.CSS_SELECTOR, 'span[class^="availability"]').get_attribute('class')
                    # if str('a-0').strip().lower() in  str(span_tag_class).strip().lower(): availability.append('Draft')
                    # if str('a-1').strip().lower() in  str(span_tag_class).strip().lower(): availability.append('Draft')
                    if str('a-2').strip().lower() in  str(span_tag_class).strip().lower(): availability.append('Active')
                    else: availability.append('Draft')
                    # elif str('a-3').strip().lower() in  str(span_tag_class).strip().lower(): availability.append('Draft')
                except: 
                    availability.append('Not Available')

            
            # for j in range(1, len(trs)):
            try:
                for button in self.browser.find_elements(By.XPATH, '//svg-icon[@class="arrow-icon"]'):
                    ActionChains(self.browser).move_to_element(button).perform()
                    button.click()
                    sleep(0.5)
                    for _ in range(0, 20):
                        try:
                            tags = self.browser.find_elements(By.XPATH, '//table[@class="table table-borderless inner-table open-shadow"]/tr[@class="d-flex drawer"]')
                            for tag in tags:    
                                g = str(tag.find_element(By.CSS_SELECTOR, 'div[class$="ean-detail"] > p > span').text).strip()
                                if g: 
                                    if g not in gtin: gtin.append(g)
                                else: gtin.append('')
                            # close_element = self.browser.find_element(By.XPATH, '//table[@class="table table-borderless inner-table open-shadow"]/tr[@class="d-flex"]').find_element(By.XPATH, '//svg-icon[@class="arrow-icon"]')
                            # ActionChains(self.browser).move_to_element(close_element).perform()
                            # close_element.click()
                            # sleep(0.3)
                            break
                        except Exception as e:pass
            except Exception as e: pass
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_size_price_status: {str(e)}')
            else: pass
        finally: return size_titles, wholesale_prices, listing_prices, availability, gtin, sizes

    def save_to_json(self, products: list[Product]):
        try:
            json_products = []
            for product in products:
                json_varinats = []
                for index, variant in enumerate(product.variants):
                    json_varinat = {
                        'position': (index + 1), 
                        'title': variant.title, 
                        'sku': variant.sku, 
                        'inventory_quantity': variant.inventory_quantity,
                        'found_status': variant.found_status,
                        'wholesale_price': variant.wholesale_price,
                        'listing_price': variant.listing_price, 
                        'barcode_or_gtin': variant.barcode_or_gtin,
                        'size': variant.size,
                        'weight': variant.weight
                    }
                    json_varinats.append(json_varinat)
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code, 
                    'frame_color': product.frame_color, 
                    'lens_code': product.lens_code, 
                    'lens_color': product.lens_color, 
                    'status': product.status, 
                    'type': product.type, 
                    'url': product.url, 
                    'metafields': [
                        { 'key': 'for_who', 'value': product.metafields.for_who },
                        { 'key': 'product_size', 'value': product.metafields.product_size }, 
                        { 'key': 'lens_material', 'value': product.metafields.lens_material }, 
                        { 'key': 'lens_technology', 'value': product.metafields.lens_technology }, 
                        { 'key': 'frame_material', 'value': product.metafields.frame_material }, 
                        { 'key': 'frame_shape', 'value': product.metafields.frame_shape },
                        { 'key': 'gtin1', 'value': product.metafields.gtin1 }, 
                        { 'key': 'img_url', 'value': product.metafields.img_url }
                    ],
                    'variants': json_varinats
                }
                json_products.append(json_product)
            
           
            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            else: pass

    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())
            products = []

            for json_d in json_data:
                number, frame_code, brand, img_url, frame_color, lens_color = '', '', '', '', '', ''
                # product = Product()
                brand = json_d['brand']
                number = str(json_d['number']).strip().upper()
                if '/' in number: number = number.replace('/', '-').strip()
                # product.name = str(json_d['name']).strip().upper()
                frame_code = str(json_d['frame_code']).strip().upper()
                if '/' in frame_code: frame_code = frame_code.replace('/', '-').strip()
                frame_color = str(json_d['frame_color']).strip().title()
                # product.lens_code = str(json_d['lens_code']).strip().upper()
                lens_color = str(json_d['lens_color']).strip().title()
                # product.status = str(json_d['status']).strip().lower()
                # product.type = str(json_d['type']).strip().title()
                # product.url = str(json_d['url']).strip()
                # metafields = Metafields()
                
                for json_metafiels in json_d['metafields']:
                    # if json_metafiels['key'] == 'for_who':metafields.for_who = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'product_size':metafields.product_size = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'activity':metafields.activity = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_material':metafields.lens_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'graduabile':metafields.graduabile = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'interest':metafields.interest = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_technology':metafields.lens_technology = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_material':metafields.frame_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_shape':metafields.frame_shape = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'gtin1':metafields.gtin1 = str(json_metafiels['value']).strip().title()
                    if json_metafiels['key'] == 'img_url':img_url = str(json_metafiels['value']).strip()
                    # elif json_metafiels['key'] == 'img_360_urls':
                    #     value = str(json_metafiels['value']).strip()
                    #     if '[' in value: value = str(value).replace('[', '').strip()
                    #     if ']' in value: value = str(value).replace(']', '').strip()
                    #     if "'" in value: value = str(value).replace("'", '').strip()
                    #     for v in value.split(','):
                    #         metafields.img_360_urls = str(v).strip()
                # product.metafields = metafields
                for json_variant in json_d['variants']:
                    sku, price = '', ''
                    # variant = Variant()
                    # variant.position = json_variant['position']
                    # variant.title = str(json_variant['title']).strip()
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    # variant.inventory_quantity = json_variant['inventory_quantity']
                    # variant.found_status = json_variant['found_status']
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    listing_price = str(json_variant['listing_price']).strip()
                    # variant.barcode_or_gtin = str(json_variant['barcode_or_gtin']).strip()
                    # variant.size = str(json_variant['size']).strip()
                    # variant.weight = str(json_variant['weight']).strip()
                    # product.variants = variant

                    image_attachment = download_image(img_url)
                    if image_attachment:
                        with open(f'Images/{sku}.jpg', 'wb') as f: f.write(image_attachment)
                        crop_downloaded_image(f'Images/{sku}.jpg')
                    data.append([number, frame_code, frame_color, lens_color, brand, sku, wholesale_price, listing_price])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                # print(response.status_code)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for downloading image')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1120
        new_height = 600
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
        elif height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')

def saving_picture_in_excel(data: list):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Model Code')
    worksheet.cell(row=1, column=2, value='Lens Code')
    worksheet.cell(row=1, column=3, value='Color Frame')
    worksheet.cell(row=1, column=4, value='Color Lens')
    worksheet.cell(row=1, column=5, value='Brand')
    worksheet.cell(row=1, column=6, value='SKU')
    worksheet.cell(row=1, column=7, value='Wholesale Price')
    worksheet.cell(row=1, column=8, value='Listing Price')
    worksheet.cell(row=1, column=9, value="Image")

    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])
        worksheet.cell(row=new_index, column=7, value=d[6])
        worksheet.cell(row=new_index, column=8, value=d[7])

        image = f'Images/{d[-3]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='I'+str(new_index))
            # col_letter = get_column_letter(7)
            # worksheet.column_dimensions[col_letter].width = width

    workbook.save('Thelios Results.xlsx')

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    # download chromedriver.exe with same version and get its path
    # if os.path.exists('chromedriver.exe'): os.remove('chromedriver.exe')
    if os.path.exists('Thelios Results.xlsx'): os.remove('Thelios Results.xlsx')

    # chromedriver_autoinstaller.install(path)
    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False
    
    f = open('Thelios start.json')
    json_data = json.loads(f.read())
    f.close()

    brands = json_data['brands']

    
    f = open('requirements/thelios.json')
    data = json.loads(f.read())
    f.close()

    store = Store()
    store.link = data['url']
    store.username = data['username']
    store.password = data['password']
    store.login_flag = True

    result_filename = 'requirements/Thelios Results.json'

    if not os.path.exists('Logs'): os.makedirs('Logs')

    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')

    scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    logs_filename = f'Logs/Logs {scrape_time}.txt'
    
    Thelios_Scraper(DEBUG, result_filename, logs_filename).controller(store, brands)
    
    for filename in glob.glob('Images/*'): os.remove(filename)
    data = read_data_from_json_file(DEBUG, result_filename)
    os.remove(result_filename)

    saving_picture_in_excel(data)
except Exception as e:
    if DEBUG: print('Exception: '+str(e))
    else: pass
